import requests
import openpyxl

def read_data(filename,sheetname):#定义函数封装
    wb = openpyxl.load_workbook(filename) #加载工作簿
    # print(wb)
    sheet = wb[sheetname]
    #获取表单
    max_row = sheet.max_row  #获取最大行数
    # print(max_row)
    case_list=[]   #创建一个空列表存放测试用例
    for i in range(2,max_row+1):
        dict1 = dict(  #数据打包成字典
        case_id = sheet.cell(row=i, column=1).value, #获取case_id
        url = sheet.cell(row=i,column=5).value,#获取Url
        data = sheet.cell(row=i,column=6).value, #获取data
        expect = sheet.cell(row=i,column=7).value,
        )
        #获取expect
        case_list.append(dict1)  # 每循环一次，把读取到的字典数据放进list中
    return case_list#返回测试用例列表

# 执行接口函数
def api_fun(url,data):
    # # 登录接口地址
    headers_reg = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}  # 请求头-字典
    # #请求正文
    res=requests. post (url=url, json=data, headers=headers_reg)
    response = res.json()
    return response

#写入结果
def write_result(filename,sheetname,row,column,final_reult):
    wb=openpyxl.load_workbook(filename)
    sheet =wb[sheetname]
    cell=sheet.cell(row=row,column=column) .value=final_reult #直接写入结果
    wb.save(filename)   #保存

#执行测试用例并回写实际结果
def execute_fun(filename,sheetname):
    cases=read_data('test_case_api.xlsx','register')    #调用读取测试用例，保存到一个变量中
    # print(cases)
    for case in cases:
        case_id=case.get('case_id')  # 取id
        url=case.get('url')
        data =eval(case.get('data'))
        expect =eval( case.get('expect'))   #获取预期结果
        expect_msg=expect.get('msg')  #获取我们预期结果中得msg信息
        real_result=api_fun(url=url,data=data)   #调用发送接口请求函数,返回结果用变量real_result接受
        real_msg=real_result.get('msg')   #获取实际结果中的msg信息
        print('预期结果中的msg:{}'.format(expect_msg))
        print('实际结果中的msg:{}'.format(real_msg))
        if real_msg==expect_msg:
            print('第{}条用例通过'.format(case_id))
            final_re='Passed'
        else:
            print('第{}条用例不通过'.format(case_id))
            final_re = 'Failed'
        write_result('test_case_api.xlsx','register',case_id+1,8,final_re)
        print('*'*25)

execute_fun('test_case_api.xlsx','login')
